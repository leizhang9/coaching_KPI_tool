from __future__ import print_function, unicode_literals
from prompt_toolkit import prompt
import argparse
import jira
import logging
from datetime import datetime, timezone, timedelta
import holidays
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

start_hour = 8 
end_hour = 17
workday_hours = (end_hour - start_hour) - 1
filename = 'vector_bugs_with_filter2.csv'
filename_summarized = 'summarized.csv'
filename_history = 'vector_bugs_with_filter2.csv'

def read_token_from_file(file_path):
    """open and read a token from a file at the specified file path.
    Args:
        file_path: The path to the file from which to read the token.
    Return:
        str: The token read from the file.
    """
    try:
        with open(file_path, 'r') as file:
            return file.read().strip()  # Remove any whitespace and newline characters
    except FileNotFoundError:
        print("Token file not found.")
        exit(1)
    except Exception as e:
        print(f"An error occurred while reading the token file: {e}")
        exit(1)

def time_string(total_hours):
    """Convert number of hours (int) into a string 'n days + m hours' """
    return f"{total_hours // workday_hours} days + {total_hours % workday_hours} hours"

def read_existing_tickets(filename):
    """read existing tickets from a CSV file and return a list of ticket keys."""
    try:
        with open(filename, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            ticket_keys = set()
            ticket_keys = {row['Ticket Key'] for row in reader}
            return ticket_keys
    except FileNotFoundError:
        return []

def write_tickets_to_csv(tickets, handover_tickets, filename):
    """
    The function ensures each ticket's data is unique and generates a detailed report including various ticket attributes, 
    processing times, statuses etc. It also reads a historical data file (filename_history) to find matching tickets from 
    the handover_tickets list and sets their 'Handed Over' status to "Yes". 
    Args:
        tickets (list): A list of ticket objects to be processed.
        handover_tickets (list): A list of ticket keys (str) that are to be marked as handed over in the csv.
                                 The function will search for these tickets in the historical data and update
                                 their status accordingly.
        filename (str): The path of the file where the CSV will be written. This file will include both updated information 
                        of active tickets and existing information of handed over tickets.
    """
    unique_tickets = {}
    # Process tickets to ensure uniqueness
    for ticket in tickets:
        unique_tickets[ticket.key] = ticket
    with open(filename_history, mode='r', newline='', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        matching_rows = []
        for row in reader:
            for ticket in handover_tickets:
                if row[0].startswith(ticket):
                    row[24] = "Yes"  # Change Handed Over column to "Yes"
                    row[12] = "Hand-over"  # Change current phase column to "Hand-over"
                    row[13] = row[13][:-1] + ", 'Hand-over']"  # Add Hand-over in phase transition
                    matching_rows.append(row)
                    break  # Break to avoid multiple matches for the same row
    with open(filename, mode='w', newline='', encoding='utf-8') as file:
        # Write the header
        fieldnames = ['Ticket Key', 'Ticket Summary', 'Issuetype', 'Assignee', 'Reporter', 'Status', 'Priority', 'Component/s',
                      'Resolution', 'Fix Version/s', 'Created Date', 'Processing Time', 'Current Phase', 'Phase Transition',
                      'Pre-analysis BMW', 'Pre-analysis Vector', 'Analysis Vector', 'Analysis BMW', 'Hand-over', 'Solution Phase',
                      'Number of cycles', 'Warning', 'Action', 'Comment', 'Handed Over', 'Category', 'Ticket Quality', 'ESCAN Status', 
                      'ESCAN_date', 'Performance Status', 'Last Comment from Vector', "Last Updated"]
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        # Write ticket data
        for ticket in unique_tickets.values():
            bmw_issue = BMWIssue(ticket)
            phase_durations, phase_list, warning_list, comment = bmw_issue.get_phase_durations()
            current_phase = phase_list[-1]
            action = []
            if current_phase in warning_list:
                if current_phase == 'Pre-analysis BMW':
                    action.append('Escalate to dispatcher/ATL') 
                elif current_phase == 'Analysis Vector':
                    action.append('Escalate to dispatcher or put on agenda for daily meeting​')
                elif current_phase == 'Analysis BMW':
                    action.append('Escalate to BMW dispatcher and put on agenda for daily meeting')
                elif current_phase == 'Solution Phase':
                    action.append('Escalate to APO or put on agenda for daily meeting')
                if 'Number of cycles' in warning_list:
                    action.append('Escalate to dispatcher/PTS and put on agenda for daily meeting​')
            ESCAN_status = "none"
            if "ESCAN_fixed" in str(ticket.fields.labels):
                ESCAN_status = "fixed"
            elif "ESCAN_open" in str(ticket.fields.labels):
                ESCAN_status = "open"
            if 'Vector_POC' in ticket.fields.labels:
                performance_status = "Vector POC"
            elif 'ipn_jc_aasr' in ticket.fields.labels:
                performance_status = "IPN JC AASR"
            else:
                performance_status = "None"

            ticket_data = {
                'Ticket Key': ticket.key,
                'Ticket Summary': ticket.fields.summary,
                'Issuetype': ticket.fields.issuetype,
                'Assignee': ticket.fields.assignee.displayName if ticket.fields.assignee else "Unassigned",
                'Reporter': ticket.fields.reporter.displayName if ticket.fields.reporter else "Unknown",
                'Status': ticket.fields.status.name,
                'Priority': ticket.fields.priority.name,
                'Component/s': ", ".join([component.name for component in ticket.fields.components]) if ticket.fields.components else "None",
                'Resolution': ticket.fields.resolution.name if ticket.fields.resolution else "Unresolved",
                'Fix Version/s': ", ".join([version.name for version in ticket.fields.fixVersions]) if ticket.fields.fixVersions else "None",
                'Created Date': ticket.fields.created,
                'Processing Time': time_string(calculate_processing_time(ticket)),
                'Current Phase': current_phase,
                'Phase Transition': phase_list,
                'Pre-analysis BMW': time_string(phase_durations['Pre-analysis BMW']) if 'Pre-analysis BMW' in phase_durations else "0",
                'Pre-analysis Vector': time_string(phase_durations['Pre-analysis Vector']) if 'Pre-analysis Vector' in phase_durations else "0",
                'Analysis Vector': time_string(phase_durations['Analysis Vector']) if 'Analysis Vector' in phase_durations else "0",
                'Analysis BMW': time_string(phase_durations['Analysis BMW']) if 'Analysis BMW' in phase_durations else "0",
                'Hand-over': time_string(phase_durations['Hand-over']) if 'Hand-over' in phase_durations else "0",
                'Solution Phase': time_string(phase_durations['Solution Phase']) if 'Solution Phase' in phase_durations else "0",
                'Number of cycles': phase_durations.get('Number of cycles', 0),
                'Warning': "There is a warning due to too long " + ", ".join(warning_list) + "." if warning_list else "None",
                'Action': ", ".join(action) + "." if action else "None",
                'Comment': ", ".join(comment) if comment else "None",
                'Handed Over': "No",
                'Category': ", ".join(bmw_issue.get_categories()) if bmw_issue.get_categories() else "None",
                'Ticket Quality': "Yes" if "ticket_quality" in ticket.fields.labels else "No",
                "ESCAN Status": ESCAN_status,
                "ESCAN_date": bmw_issue.get_escan_date(),
                'Performance Status': 'None' if not performance_status else performance_status,
                'Last Comment from Vector': get_last_comment_date_from_vector(ticket),
                'Last Updated': datetime.strptime(ticket.fields.updated, "%Y-%m-%dT%H:%M:%S.%f%z").date()
            }
            writer.writerow(ticket_data)
        # Write handed over tickets
        writer = csv.writer(file)
        writer.writerows(matching_rows)

def write_tickets_to_csv_summarized(tickets, filename):
    """Writes a summarized report of ticket data to a CSV file """
    unique_tickets = {}
    # Process tickets to ensure uniqueness
    for ticket in tickets:
        unique_tickets[ticket.key] = ticket
    with open(filename, mode='w', newline='', encoding='utf-8') as file:
        # Write the header
        fieldnames = ['Ticket Key', 'Ticket Summary', 'Issuetype', 'Assignee', 'Priority', 'Current Phase', 
                      'Current Phase Duration', 'Warning', 'Reopened', 'Handed Over', 'Category', 'Last Comment from Vector']
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        # Write ticket data
        for ticket in unique_tickets.values():
            bmw_issue = BMWIssue(ticket)
            phase_durations, phase_list, warning_list, comment = bmw_issue.get_phase_durations()
            current_phase = phase_list[-1]

            ticket_data = {
                'Ticket Key': ticket.key,
                'Ticket Summary': ticket.fields.summary,
                'Issuetype': ticket.fields.issuetype,
                'Assignee': ticket.fields.assignee.displayName if ticket.fields.assignee else "Unassigned",
                'Priority': ticket.fields.priority.name,
                'Current Phase': current_phase,
                'Current Phase Duration': time_string(phase_durations[current_phase]),
                'Warning': "There is a warning due to too long " + ", ".join(warning_list) + "." if warning_list else "None",
                'Reopened': "Yes" if "Ticket is reopened" in comment else "No",
                "Handed Over": "No",
                'Category': ", ".join(bmw_issue.get_categories()) if bmw_issue.get_categories() else "None ",
                'Last Comment from Vector': get_last_comment_date_from_vector(ticket)
            }
            writer.writerow(ticket_data)

def calculate_work_hours(start, end):
    """calculate working hours between two dates"""
    my_holidays = holidays.CountryHoliday('DE')
    total_hours = 0
    current = start
    while current < end:
        if start_hour <= current.hour < end_hour and current.weekday() < 5 and current.date() not in my_holidays:
            total_hours += 1
        current += timedelta(hours=1)
    return total_hours

def calculate_processing_time(ticket):
    """Calculate processing time in hours"""
    # detect if a ticket starts as a Bug and then changes to a TAA Defect
    flag_transitioned = False
    for history in ticket.changelog.histories:
        for change in history.items:
            if change.toString is not None and "TAA Defect" in change.toString:
                # A transition to a TAA Defect is detected
                flag_transitioned = True
                print("A transition to a TAA Defect is detected")
                change_date = datetime.strptime(history.created, "%Y-%m-%dT%H:%M:%S.%f%z")
                break
        if flag_transitioned:
            break

    now = datetime.now(timezone.utc)  # 2024-04-07 20:04:23.350658+00:00

    created_date = datetime.strptime(ticket.fields.created, "%Y-%m-%dT%H:%M:%S.%f%z") if not flag_transitioned else change_date
    if (ticket.fields.status.name == "Resolved") or (ticket.fields.status.name == "Closed"):
        resolved_date = datetime.strptime(ticket.fields.resolutiondate, "%Y-%m-%dT%H:%M:%S.%f%z")
        total_hours = calculate_work_hours(created_date, resolved_date)
    else:
        total_hours = calculate_work_hours(created_date, now)
    return total_hours

def get_last_comment_date_from_vector(ticket, vector_team_file='vector_team.txt'):
    """Gets the date of the last comment made by a member of the Vector team on a given ticket."""
    comments = ticket.fields.comment.comments
    if comments:
        for comment in reversed(comments):
            last_comment_date = datetime.strptime(comment.created, "%Y-%m-%dT%H:%M:%S.%f%z").date()
            last_comment_author = comment.author.displayName if comment.author else "Unknown"
            is_vector_member = is_vector_team_member(last_comment_author, vector_team_file)
            if is_vector_member:
                return last_comment_date
    return "No comments by Vector members found"

def get_last_comment_by_vector(ticket, vector_team_file='vector_team.txt'):
    """Gets the content of the last comment made by a member of the Vector team on a given ticket."""
    comments = ticket.fields.comment.comments
    # Iterate over comments from the latest to the earliest
    for comment in reversed(comments):
        last_comment_author = comment.author.displayName if comment.author else "Unknown"
        is_vector_member = is_vector_team_member(last_comment_author, vector_team_file)
        if is_vector_member:
            return comment.body
    return "No comments by Vector members found."

def is_vector_team_member(assignee, vector_team_file='vector_team.txt'):
    """Checks if the assignee is a member of the Vector team."""
    try:
        with open(vector_team_file, 'r') as file:
            vector_team_members = [line.strip() for line in file.readlines()]
            return assignee in vector_team_members
    except FileNotFoundError:
        print("Vector team file not found.")
        return False
    
def get_ticket_phase(status, labels, assignee, fixVersions, issuetype, components):
    """Determine the current phase of the ticket"""
    # print(status, labels, is_vector_team_member(assignee), assignee, components)
    if status in ["Resolved", "Integrated", "Closed", "Following"]:
        return "Resolved"
    if status == "New" and not is_vector_team_member(assignee) and "AA_Stack" in str(components) and "Vector_relevant" in str(labels):
        return "Pre-analysis BMW"
    if status == "New" and "AA_Stack" in str(components) and "Vector_relevant" in str(labels):
        return "Pre-analysis Vector"
    if status == "Open" and "Vector_relevant" in str(labels) and is_vector_team_member(assignee) and "AA_Stack" in str(components) and not("vector_wait_retest" in str(labels) or "Vector_Blocked" in str(labels)):
        return "Analysis Vector"
    if issuetype == "Bug":
        if (status == "Pending" ) and "AA_Stack" in str(components) and "Vector_relevant" in str(labels): 
            return "Analysis BMW"
    elif issuetype == "TAEE Defect":
        if ("vector_wait_retest" in str(labels) or "Vector_Blocked" in str(labels)) and "AA_Stack" in str(components) and "Vector_relevant" in str(labels) and (status in ["Open", "New"] ):
            return "Analysis BMW"
    if (status == "In Progress" or status == "In Test" )and fixVersions is not None and "AA_Stack" in str(components) and "Vector_relevant" in str(labels):
        return "Solution Phase"
    if (status == "New" or status == "Open" or status == "Pending" or status == "In Progress" or status == "In Test") and not("AA_Stack" in str(components) and "Vector_relevant" in str(labels)):
        return "Hand-over"
    return "Unknown"


class BMWIssue:
    def __init__(self,  issue):
        self.issue = issue
        self.histories = self.issue.changelog.histories if issue.changelog else []
        self.initial_status = "New"
        self.initial_labels = "Unknown"
        self.initial_assignee = "Unassigned"
        self.initial_fixVersions = []
        self.initial_issuetype = "Bug"
        self.priority = self.issue.fields.priority.name
        self.threshold = {'Pre-analysis BMW': {'Top': 1 * workday_hours, 'Critical': 1 * workday_hours,'High': 1 * workday_hours, 'Medium': 3 * workday_hours, 'Low': float('inf'), 'None': float('inf')},
                        'Pre-analysis Vector': {'Top': 1 * workday_hours, 'Critical': 1 * workday_hours,'High': 1 * workday_hours, 'Medium': 3 * workday_hours, 'Low': float('inf'), 'None': float('inf')},
                        'Analysis Vector': {'Top': 2 * workday_hours, 'Critical': 4 * workday_hours,'High': 4 * workday_hours, 'Medium': 8 * workday_hours, 'Low': float('inf'), 'None': float('inf')},
                        'Analysis BMW': {'Top': 2 * workday_hours, 'Critical': 2 * workday_hours,'High': 2 * workday_hours, 'Medium': 2 * workday_hours, 'Low': float('inf'), 'None': float('inf')},
                        'Hand-over': {'Top': float('inf'), 'Critical': float('inf'),'High': float('inf'), 'Medium': float('inf'), 'Low': float('inf'), 'None': float('inf')},
                        'Solution Phase': {'Top': 3 * workday_hours, 'Critical': 3 * workday_hours,'High': 3 * workday_hours, 'Medium': 3 * workday_hours, 'Low': float('inf'), 'None': float('inf')},
                        'Resolved': {'Top': float('inf'), 'Critical': float('inf'),'High': float('inf'), 'Medium': float('inf'), 'Low': float('inf'), 'None': float('inf')},
                        'Processing time': {'Top': 3 * workday_hours, 'Critical': 5 * workday_hours,'High': 6 * workday_hours, 'Medium': 10 * workday_hours, 'Low': 10 * workday_hours, 'None': float('inf')},
                        'Unknown': {'Top': 0 * workday_hours, 'Critical': 0 * workday_hours,'High': 0 * workday_hours, 'Medium': 0 * workday_hours, 'Low': float('inf'), 'None': float('inf')},
                        'Number of cycles': {'Top': 2, 'Critical': 2,'High': 2, 'Medium': 2, 'Low': float('inf'), 'None': float('inf')}}

    #return all the histories. In a list
    def get_changelog_histories(self):
        return self.changelog.histories

    #return changes inside a history in a list
    def get_history_changes(story):
        return story.items

    def get_change_type(change):
        return change.field 
    
    def was_vector_relevant(self):
        #first get all the labels changes
        label_changes=[]
        for story in self.histories:
            for change in story.items:
                if change.field == "labels":
                    if "Vector_relevant" in change.toString:
                        return True
        return False
    
    def is_vector_relevant(self):
        return "Vector_relevant" in self.issue.fields.labels

    @staticmethod
    def date_from_history(history):
        return datetime.strptime(history.created, "%Y-%m-%dT%H:%M:%S.%f%z")

    def get_fixed_version_changes(self):
        for history in self.histories:
            for change in history.items:
                if "Fix Version" in change.field:
                    print("history on date: ", self.date_from_history(history))
                    print("\tchange.field: ", change.field, "change.fromString:", change.fromString, "change.toString:", change.toString)

    def get_phase_durations(self):
        """
        Calculates the duration of various phases of a ticket based on its change history. It tracks shifts between phases and stores in 'phase_list'.
        It also tracks the number of cycles (cyclic transitions between Analysis Vector and Analysis BMW). The function also assesses warnings based on 
        threshold durations that could suggest a need for escalation based on ticket priority. 'comment' stores corner cases where a ticket enters an undefined phase.
        Args: 
            None
        Returns: 
            phase_durations (dict): A dictionary where keys are phase names (str) and values are the accumulated duration (int) in hours spent in each phase.
            phase_list (list): A list capturing the sequence of phases the ticket has gone through.
            warning_list (list): A list of strings indicating phases where the duration has exceeded predefined thresholds.
            comment (set): A set of strings noting corner cases where a ticket enters an undefined phase.
        """
        phase_durations = {"Pre-analysis BMW": 0, "Pre-analysis Vector": 0, "Analysis Vector": 0, "Analysis BMW": 0, "Hand-over": 0, "Solution Phase": 0, "Resolved": 0, "Unknown": 0}
        phase_durations['Number of cycles'] = 0
        phase_list =  []
        comment = set()
        status = self.initial_status
        labels = self.initial_labels
        assignee = self.initial_assignee
        fixVersions = self.initial_fixVersions
        issuetype = self.initial_issuetype
        components = self.issue.fields.components

        # Initialize the phase and end_date
        old_phase = "Pre-analysis BMW"
        new_phase = old_phase
        end_date = datetime.strptime(self.issue.fields.created, "%Y-%m-%dT%H:%M:%S.%f%z")
        flag_half_cycle = False
        flag_full_cycle = False
        # for debug
        print( "\t", self.issue.key)

        for history in self.histories:  # Oldest first
            temp_status = status
            temp_labels = labels
            temp_assignee = assignee
            temp_fixVersions = fixVersions
            temp_issuetype = issuetype
            temp_components = components
            # determine the phase at the time of change
            for change in history.items:
                # for debug
                if change.field == "status" or change.field == "labels" or change.field == "assignee" or "Fix Version" in change.field:
                     print("history on date:", self.date_from_history(history))
                     print("\tchange.field:", change.field, "change.fromString:", change.fromString, "change.toString:", change.toString)

                flag_change = False
                if change.field == "status":
                    temp_status = change.toString if change.toString else temp_status
                    flag_change = True
                    if change.toString == "Resolved" and new_phase !="Solution Phase":
                        comment.add('Status is set to "Resolved" before setting to "In Progress" during ' + new_phase)
                    # mark reopened tickets
                    if change.toString == "Open" and change.fromString != "New":
                        comment.add("Ticket is reopened")
                elif change.field == "labels":
                    temp_labels = change.toString if change.toString else temp_labels
                    flag_change = True
                elif change.field == "assignee":
                    temp_assignee = change.toString if change.toString else temp_assignee
                    flag_change = True
                elif "Fix Version" in change.field:
                    temp_fixVersions = [change.toString] if change.toString else temp_fixVersions
                    flag_change = True
                elif change.field == "issuetype":
                    temp_issuetype = change.toString if change.toString else temp_issuetype
                    flag_change = True
                elif change.field == "components":
                    temp_components = change.toString if change.toString else temp_components
                    flag_change = True
                # Get number of cycles
                if (change.fromString is not None and change.toString is not None) and \
                    ((change.fromString == "Open" and change.toString == "Pending") or ("vector_wait_retest" in change.toString or "Vector_Blocked" in change.toString)) and not flag_half_cycle:
                    flag_half_cycle = True
                if (change.fromString is not None and change.toString is not None) and \
                    (change.fromString == "Pending" and change.toString == "Open") and is_vector_team_member(temp_assignee) and flag_half_cycle:
                    flag_full_cycle = True

            status, labels, assignee, fixVersions, issuetype, components = temp_status, temp_labels, temp_assignee, temp_fixVersions, temp_issuetype, temp_components
            if flag_change:
                new_phase = get_ticket_phase(status, labels, assignee, fixVersions, issuetype, components)
            # "Hand-over" only after "Analysis Vector"
            if new_phase == "Hand-over" and phase_list and phase_list[-1] != "Analysis Vector":
                new_phase == old_phase
            # Fix the corner case
            if new_phase == "Unknown":
                if old_phase == "Pre-analysis BMW":
                    new_phase = old_phase
                    comment.add("During Pre-analysis BMW, status is changed before being assigned to Vector")
                elif old_phase == "Pre-analysis Vector":
                    new_phase = old_phase
                    comment.add("Corner case during Pre-analysis Vector")
                elif old_phase == "Analysis Vector" and (change.field == "assignee" and not is_vector_team_member(change.toString)):
                    new_phase = old_phase
                    comment.add("During Analysis Vector, assignee is changed to non-Vector member without changing status or label")
                elif old_phase == "Analysis Vector" and "During Analysis Vector, assignee is changed to non-Vector member without changing status or label" not in comment:
                    new_phase = old_phase
                    comment.add("Corner case during Analysis Vector")
                elif old_phase == "Solution Phase":
                    new_phase = old_phase
                    comment.add("During Solution Phase, status is changed")
                elif old_phase == "Analysis BMW":
                    new_phase = old_phase
                    comment.add("Corner case during Analysis BMW")
                else:
                    new_phase = old_phase

            if new_phase != old_phase:  # phase change
                phase_list.append(old_phase)
                # for debug
                print("\tphase change from:", old_phase, " to:", new_phase, 'date:', self.date_from_history(history))

                begin_date = end_date
                end_date = self.date_from_history(history)
                duration = calculate_work_hours(begin_date, end_date)
                phase_durations[old_phase] = phase_durations[old_phase] + duration
                old_phase = new_phase

                if flag_full_cycle:
                    phase_durations['Number of cycles'] += 1
                    flag_half_cycle = False
                    flag_full_cycle = False

                # add comment when change to pending without setting label
                if new_phase == "Analysis BMW" and ("Change to pending without setting label" not in comment) and not("vector_wait_retest" in labels or "Vector_Blocked" in labels):
                    comment.add("Change to pending without setting label")
                    
        now = datetime.now(timezone.utc)
        begin_date = end_date
        current_phase = get_ticket_phase(str(self.issue.fields.status.name), str(self.issue.fields.labels), str(self.issue.fields.assignee.displayName), 
                                         str(self.issue.fields.fixVersions), str(self.issue.fields.issuetype.name), str(self.issue.fields.components))
        # Fix corner case
        if current_phase == "Unknown":
                if new_phase == "Pre-analysis BMW":
                    current_phase = new_phase
                    comment.add("During Pre-analysis BMW, status was changed before being assigned to Vector")
                elif new_phase == "Pre-analysis Vector":
                    current_phase = new_phase
                    comment.add("Corner case during Pre-analysis Vector")
                elif new_phase == "Analysis Vector" and (change.field == "assignee" and not is_vector_team_member(change.toString)):
                    current_phase = new_phase
                    comment.add("During Analysis Vector, assignee is changed to non-Vector member without changing status or label")
                elif new_phase == "Analysis Vector":
                    current_phase = new_phase
                    comment.add("Corner case during Analysis Vector")
                elif new_phase == "Solution Phase":
                    current_phase = new_phase
                    comment.add("Corner case during Solution Phase")
                elif new_phase == "Analysis BMW":
                    current_phase = new_phase
                    comment.add("Corner case during Analysis BMW")
                elif new_phase == "Hand-over":
                    current_phase = new_phase
                    comment.add("Corner case during Hand-over")
                elif new_phase == "Resolved":
                    current_phase = new_phase
                    comment.add("Corner case during Resolved")

        current_duration = calculate_work_hours(begin_date, now)
        phase_durations[current_phase] = phase_durations[current_phase] + current_duration
        if not phase_list:
            phase_list.append("Pre-Analysis BMW")
        if current_phase != phase_list[-1]:
            phase_list.append(current_phase)
        # get warning
        warning_list = []
        for phase in phase_durations.keys():
            if phase_durations[phase] > self.threshold[phase][self.priority]:
                warning_list.append(phase)
        # Orange color as potential candidate for escalation for tickets that reach 50% of the Analysis Vector threshold.
        if current_phase == "Analysis Vector" and phase_durations["Analysis Vector"] * 2 >= self.threshold["Analysis Vector"][self.priority] \
        and phase_durations["Analysis Vector"] <= self.threshold["Analysis Vector"][self.priority]:
            warning_list.append("reaching 50%% of the Analysis Vector threshold")
        if calculate_processing_time(self.issue) > self.threshold['Processing time'][self.priority]:
            warning_list.append('Processing time')
        # Warning if last updated for high and critical priority is longer than 1day
        if (self.priority == 'High') or (self.priority == 'Critical'):
            if calculate_work_hours(datetime.strptime(self.issue.fields.updated, "%Y-%m-%dT%H:%M:%S.%f%z"), now) > 9:
                warning_list.append('time since last updated')
                # print("too long time since last updated: ", calculate_work_hours(datetime.strptime(self.issue.fields.updated, "%Y-%m-%dT%H:%M:%S.%f%z"), now))
        return phase_durations, phase_list, warning_list, comment
    
    def get_categories(self):
        categories = set()
        comments = self.issue.fields.comment.comments  # Get all comments for the ticket
        category_labels = ['ipsec', 'diag_config', 'diag_tester', 'diag_req', 'e2e', 'shutdown', 'someip_peer', 'someip_config', 
                          'em_timeout', 'secpol', 'user_callback', 'amsr', 'not_amsr', 'doc']
        for comment in comments:
            # print(comment.body)
            for category_label in category_labels:
                if 'category: ' + category_label in comment.body.lower():
                    categories.add(category_label)
        return categories
    
    def get_escan_date(self):
        for history in self.histories:  # Oldest first
            for change in history.items:
                if change.field == "labels" and "ESCAN_open" in change.toString:
                    return datetime.strptime(history.created, "%Y-%m-%dT%H:%M:%S.%f%z").date()
        return "none"

        
class AAStackJira:
    def __init__(self, user):
        self.cc_url = 'https://jira.cc.bmwgroup.net/'
        self.cc_jira = jira.JIRA(self.cc_url, token_auth=token)

    def get_vector_bugs(self):
        # vector_bugs = self.cc_jira.search_issues('project = "Software Platform" and labels in (Vector_relevant) and created >= "2023-11-01"', expand='changelog')
        vector_bugs = set()

        #vector_bugs = self.cc_jira.search_issues('project = SWP AND issuetype in (Task, Bug, "TAEE Defect") AND labels in (ESCAN_fixed, ESCAN_open, ESCAN_planned) AND Variant in (ipn-10)', \
        #                                        expand='changelog')
        #vector_bugs = self.cc_jira.search_issues('type in (Bug, "TAEE Defect") AND project = SWP AND assignee was in (hoomanhahabibipartner, praveenkumardarapureddypartner, keerthikumarpartner, sebastiandoerrpartner, \
        #                                         abinanallusamysathiyamoorthypartner, saaddadebopartner, mahmoudelshrbienypartner, harishbudereddypartner) AND Variant in (ipn-10)',\
         #                                        expand='changelog', maxResults = 500)
        # for test
        # vector_bugs = self.cc_jira.search_issues('key = SWP-107093', expand='changelog')
        ##test without performance
        #vector_bugs = self.cc_jira.search_issues('filter = 42437 AND (labels in (Vector_relevant) OR assignee = michaelmoeckpartner ) AND status != Following \
        #                                         AND labels not in (ipn_jc_aasr) ORDER BY issuetype DESC, priority DESC', \
        #                                         expand='changelog', maxResults = 500)
        # for performance no POC
        vector_bugs = self.cc_jira.search_issues('filter = 42437 AND (labels in (Vector_relevant) OR assignee = michaelmoeckpartner) AND status != Following \
                                                AND labels not in (Vector_POC) ORDER BY priority DESC, issuetype DESC', \
                                                expand='changelog', maxResults = 500)

        expanded = False
        if expanded is True:
            vector_bugs = self.cc_jira.search_issues('filter = 42437 and (labels in (Vector_relevant) OR assignee = michaelmoeckpartner) \
                                                and status != Following \
                                                and project in (SWP, IPNDEV, "IPN UserSpaceLibs") AND "Variant 2" in (IPN-10_PERF, IPN-10_MAIN) \
                                                and issuetype in (Bug, "TAEE Defect") AND status not in (closed, following, resolved) \
                                                and (affectedVersion in versionMatch("^25-07*") OR affectedVersion in versionMatch("^13.*")) \
                                                AND component = AA_Stack  AND project = SWP AND (labels in (Vector_relevant) OR assignee = michaelmoeckpartner ) \
                                                AND status != Following \
                                                and labels not in (ipn_performance_daily) ORDER BY issuetype DESC, priority DESC ', \
                                                expand='changelog')
          
        csv_tickets_key = read_existing_tickets(filename)
        print("number of tickets in csv: ", len(csv_tickets_key))
        # debug, to add more history tickets
        csv_tickets_key1 = read_existing_tickets(filename_history)
        print("number of tickets in history csv: ", len(csv_tickets_key1))
        csv_tickets_key.update(csv_tickets_key1)
        print("number of tickets in csv: ", len(csv_tickets_key))

        handover_tickets = set()
        for key in csv_tickets_key:
            filter_existing_ticket = f'key = {key}'
            csv_ticket = self.cc_jira.search_issues(filter_existing_ticket, expand='changelog')
            if not csv_ticket:
                print("this ticket cannot be found:", filter_existing_ticket)
                handover_tickets.add(key)
            else:
                vector_bugs.extend(csv_ticket)
        print("handover_tickets:", handover_tickets)
        return vector_bugs, handover_tickets

if __name__== "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-v", "--verbose", action="count", default=0, help="increase output verbosity")
    parser.add_argument("-u", "--user", help="ASCENT user", dest="user")
    #parser.add_argument("-p", "--password", help="ASCENT password", dest="password")
    parser.add_argument("-P", "--ask-password", help="ask for the ASCENT password", action="store_true")
    parser.add_argument("-t", "--token-file", help="Path to the token file", dest="token_file")
    args = parser.parse_args()

    logging.basicConfig(level=30 - args.verbose * 10)
    logging.debug(args)
    logging.debug(args.user)

    if args.ask_password:
        answers = prompt("ASCENT pasword: ", is_password=True)
        args.password = answers

    if args.token_file:
        token = read_token_from_file(args.token_file)
    else:
        print("No token file provided.")
        exit(1)

    aas_jira = AAStackJira(args.user)

    ### vector bugs
    bugs, handover_tickets = aas_jira.get_vector_bugs()

    write_tickets_to_csv(bugs, handover_tickets, filename)
    write_tickets_to_csv_summarized(bugs, filename_summarized)



    # Read the CSV file into a DataFrame
    df = pd.read_csv(filename_summarized)

    # Store 'Warning' column in a separate variable and then drop it from DataFrame before writing to Excel
    warning_data = df['Warning'].copy()
    df.drop('Warning', axis=1, inplace=True)
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define the style for text wrapping and vertical top alignment
    wrap_text = Alignment(vertical='top', wrapText=True)
    grey_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    white_border = Border(
        left=Side(border_style="thin", color="FFFFFF"),
        right=Side(border_style="thin", color="FFFFFF"),
        top=Side(border_style="thin", color="FFFFFF"),
        bottom=Side(border_style="thin", color="FFFFFF")
    )

    # Append the DataFrame rows to Excel
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(r)
        ws.row_dimensions[r_idx].height = 30 
        for cell in ws[r_idx]:
            cell.alignment = wrap_text
            cell.fill = grey_fill
            cell.border = white_border

        if r_idx == 1:  # Apply styles to header
            for cell in ws[r_idx]:
                cell.font = Font(bold=True)
                cell.fill = grey_fill
                cell.alignment = Alignment(vertical='top', horizontal="center")
        else:  # Apply hyperlink to the first column cells except for the header
            first_col_value = ws.cell(row=r_idx, column=1).value
            hyperlink = f"https://jira.cc.bmwgroup.net/browse/{first_col_value}"
            ws.cell(row=r_idx, column=1).hyperlink = hyperlink
            ws.cell(row=r_idx, column=1).font = Font(color="0000FF", underline="single")

    # Define column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 40

    # Apply conditional formatting using stored warning data
    red_font = Font(color="FF0000")
    green_font = Font(color="006400")
    orange_font = Font(color="FFA500")
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=0):
        warning_cell_value = warning_data[idx]
        current_phase_cell = row[5]
        duration_cell = row[6]
        #print(warning_cell_value, "current phase", current_phase_cell.value)
        if isinstance(warning_cell_value, str) and "There is a warning" in str(warning_cell_value) and str(current_phase_cell.value) in str(warning_cell_value) \
            and "reaching 50%% of the Analysis Vector threshold" not in warning_cell_value:
            duration_cell.font = red_font
        # Orange color as potential candidate for escalation for tickets that reach 50% of the Analysis Vector threshold.
        elif isinstance(warning_cell_value, str) and "reaching 50%% of the Analysis Vector threshold" in warning_cell_value:
            duration_cell.font = orange_font
        else:
            duration_cell.font = green_font

    # Apply autofilter to the first row
    ws.auto_filter.ref = ws.dimensions

    # Save the workbook
    wb.save('formatted_summarized.xlsx')